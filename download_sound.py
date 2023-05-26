import os
import urllib.request
from openpyxl import load_workbook  # 用于读取Excel文件


class youdao():
    def __init__(self, type=0, word='hellow'):
        word = word.lower()  # 小写
        self._type = type  # 发音方式
        self._word = word  # 单词

        # 文件根目录
        self._dirRoot = os.path.dirname(os.path.abspath(__file__))
        if 0 == self._type:
            self._dirSpeech = os.path.join(self._dirRoot, 'Sounds/Speech_US')  # 美音库
        else:
            self._dirSpeech = os.path.join(self._dirRoot, 'Sounds/Speech_EN')  # 英音库

        # 判断是否存在美音库
        if not os.path.exists('Sounds/Speech_US'):
            os.makedirs('Sounds/Speech_US')  # 不存在，就创建
        # 判断是否存在英音库
        if not os.path.exists('Sounds/Speech_EN'):
            os.makedirs('Sounds/Speech_EN')  # 不存在，就创建

    def setAccent(self, type=1):
        self._type = type  # 发音方式
        if 0 == self._type:
            self._dirSpeech = os.path.join(self._dirRoot, 'Sounds/Speech_US')  # 美音库
        else:
            self._dirSpeech = os.path.join(self._dirRoot, 'Sounds/Speech_EN')  # 英音库

    def getAccent(self):
        return self._type  # 返回发音方式

    def down(self, word):
        word = word.lower()  # 小写
        tmp = self._getWordMp3FilePath(word)
        if tmp is None:
            self._getURL()  # 组合URL
            urllib.request.urlretrieve(self._url, filename=self._filePath)  # 下载到目标地址
            print('%s.mp3 下载完成' % self._word)
        else:
            print('已经存在 %s.mp3, 不需要下载' % self._word)

    def _getURL(self):
        self._url = r'http://dict.youdao.com/dictvoice?type=' + str(
            self._type) + r'&audio=' + self._word  # 生成发音的目标URL

    def _getWordMp3FilePath(self, word):
        word = word.lower()  # 小写
        self._word = word
        self._fileName = self._word + '.mp3'
        self._filePath = os.path.join(self._dirSpeech, self._fileName)

        # 判断是否存在这个MP3文件
        if os.path.exists(self._filePath):
            return self._filePath  # 存在这个mp3
        else:
            return None  # 不存在这个MP3，返回none


if __name__ == "__main__":
    file = 'wordtest.xlsx'  # 将文件类型从 .txt 更改为 .xlsx
    wb = load_workbook(filename=file)  # 使用openpyxl打开excel文件
    sheet = wb.active  # 获取当前活动的sheet

        # 遍历每一行的数据
    for row in sheet.iter_rows(min_row=0,values_only=True):  # 从第二行开始迭代每一行，因为通常第一行是表头
        word = row[0]  # 第一列是英文单词
        sp = youdao()  # 创建有道对象
        sp.down(word)  # 下载单词的读音

